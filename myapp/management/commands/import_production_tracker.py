"""
Parse .kitchen_ops/Kitchen Operations/Production Tracker.xlsx into
historical dish-waste statistics.

Per project_popularity_learning.md, this data MUST NOT be retroactively
normalized into the live MealService schema (units are chaotic:
integers, weights "5# chicken", containers "1 Hotel Pan", ingredient lists).
This command extracts what's cleanly extractable: rows where both Prep
and Waste are numeric. Aggregates by dish and computes waste-ratio stats.

Output: .historical_stats/production_tracker.json  (gitignored)
        rendered by the /historical/ view as demo-day prop.

Run:
    python manage.py import_production_tracker
    python manage.py import_production_tracker --apply  # writes JSON
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError


# Column layout verified 2026-04-19 against sample sheets.
# (1-indexed columns per openpyxl convention)
MEAL_COLUMNS = [
    # (slot, name_col, prep_col, waste_col, usage_col)
    ('cold_breakfast', 1, 3, 5, 7),     # First Breakfast column group; we'll map both
    ('hot_breakfast',  1, 3, 5, 7),     # to generic "breakfast" in the output since the
                                         # xlsx doesn't distinguish cold vs hot
    ('lunch',          9, 11, 13, 15),
    ('dinner',         17, 19, 21, 23),
]

# We collapse breakfast (cold + hot) into a single 'breakfast' bucket for
# historical analysis, since the xlsx column layout doesn't split them.
COL_GROUPS = [
    ('breakfast', 1, 3, 5, 7),
    ('lunch',     9, 11, 13, 15),
    ('dinner',    17, 19, 21, 23),
]

POPULATED_SHEET_PATTERN = re.compile(r'^Jan\s+\d+\s+-\s+', re.I)


_NUMBER_RE = re.compile(r'^\s*(-?\d+(?:\.\d+)?)\s*$')
# Matches a leading decimal number, optional fraction, optional trailing unit.
# Examples: "8", "8 Ea", "1/4 Bag", "1.5 pk", ".25 Hotel Pan", "3/4 Bag"
_NUM_UNIT_RE = re.compile(
    r'^\s*'
    r'(?:(\d+)\s*/\s*(\d+)'          # group 1/2: N/M fraction
    r'|(-?\d+(?:\.\d+)?|\.\d+))'      # OR group 3: decimal
    r'\s*([A-Za-z#][A-Za-z #.]*)?\s*$',   # optional unit (allowing "Hotel Pan", "Ea", "#", etc.)
)


def _parse_quantity(cell) -> tuple[float | None, str]:
    """Parse a cell into (numeric_value, normalized_unit).

    Returns (None, '') if not parseable. Unit is lowercased + trimmed.
    Examples:
      8            → (8.0, '')
      '8'          → (8.0, '')
      '8 Ea'       → (8.0, 'ea')
      '1/4 Bag'    → (0.25, 'bag')
      '1.5 pk'     → (1.5, 'pk')
      '1 Hotel Pan'→ (1.0, 'hotel pan')
      '.25 Bag'    → (0.25, 'bag')
    """
    if cell is None:
        return None, ''
    if isinstance(cell, (int, float)):
        return float(cell), ''
    s = str(cell).strip()
    if not s:
        return None, ''
    m = _NUM_UNIT_RE.match(s)
    if not m:
        return None, ''
    if m.group(1) and m.group(2):
        num = float(m.group(1)) / float(m.group(2))
    else:
        try:
            num = float(m.group(3))
        except (TypeError, ValueError):
            return None, ''
    unit = (m.group(4) or '').strip().lower()
    return num, unit


def _parse_number(cell):
    """Legacy helper — just the numeric part. Use _parse_quantity for unit-aware."""
    num, _unit = _parse_quantity(cell)
    return num


def _normalize_name(name) -> str:
    """Lowercase, strip punctuation, collapse spaces. Handles Sean's typos
    by doing rough normalization."""
    s = str(name or '').lower().strip()
    # Strip trailing parenthetical notes like "(Medical Request)"
    s = re.sub(r'\s*\([^)]*\)\s*$', '', s)
    s = re.sub(r'[^a-z0-9 ]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


class Command(BaseCommand):
    help = 'Parse Production Tracker.xlsx into historical dish waste stats (JSON).'

    def add_arguments(self, parser):
        parser.add_argument('--apply', action='store_true',
                            help='Write the JSON output (default is dry-run: prints summary only).')
        parser.add_argument('--xlsx', type=str,
                            default='.kitchen_ops/Kitchen Operations/Production Tracker.xlsx',
                            help='Path to the xlsx (relative to BASE_DIR).')

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl not installed. `pip install openpyxl`.')

        xlsx_path = Path(settings.BASE_DIR) / opts['xlsx']
        if not xlsx_path.exists():
            raise CommandError(f'xlsx not found: {xlsx_path}')

        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

        # Per-service records: one per (sheet, row-block, meal-column) with
        # both numeric prep and numeric waste.
        records = []
        skipped_text_only = 0
        skipped_no_name = 0

        for sheet_name in wb.sheetnames:
            if not POPULATED_SHEET_PATTERN.match(sheet_name):
                continue
            ws = wb[sheet_name]

            for row_idx in range(1, ws.max_row + 1):
                for slot, name_col, prep_col, waste_col, usage_col in COL_GROUPS:
                    name_cell = ws.cell(row=row_idx, column=name_col).value
                    prep_cell = ws.cell(row=row_idx, column=prep_col).value
                    waste_cell = ws.cell(row=row_idx, column=waste_col).value

                    # Skip header rows
                    if name_cell in ('Product name', 'Breakfast', 'Lunch', 'Dinner', None):
                        continue
                    # Skip date markers (first column may hold a datetime or 'Jan 7,2026')
                    if isinstance(name_cell, datetime):
                        continue
                    if isinstance(name_cell, str) and re.match(r'^(jan|feb|mar)\s+\d+', name_cell.lower()):
                        continue

                    name_norm = _normalize_name(name_cell)
                    if not name_norm or len(name_norm) < 3:
                        skipped_no_name += 1
                        continue

                    prep, prep_unit = _parse_quantity(prep_cell)
                    waste, waste_unit = _parse_quantity(waste_cell)
                    if prep is None or waste is None:
                        skipped_text_only += 1
                        continue
                    if prep <= 0:
                        skipped_text_only += 1
                        continue
                    # Both values must use the same unit (or both unitless) for
                    # the ratio to be meaningful.
                    if prep_unit != waste_unit:
                        # Common case: "1 Bag" prep, ".25 Bag" waste — same unit, valid.
                        # "24" prep, "6 Ea" waste — unit mismatch but probably same thing.
                        # Allow if one side is unitless.
                        if prep_unit and waste_unit:
                            skipped_text_only += 1
                            continue

                    waste_ratio = waste / prep
                    records.append({
                        'sheet': sheet_name,
                        'row': row_idx,
                        'slot': slot,
                        'name_original': str(name_cell).strip(),
                        'name': name_norm,
                        'prep': prep,
                        'waste': waste,
                        'unit': prep_unit or waste_unit or '',
                        'waste_ratio': round(waste_ratio, 3),
                    })

        # Aggregate by normalized dish name
        by_dish: dict = defaultdict(lambda: {
            'name_original_samples': set(),
            'services': [],
            'total_prep': 0.0,
            'total_waste': 0.0,
        })
        for rec in records:
            g = by_dish[rec['name']]
            g['name_original_samples'].add(rec['name_original'])
            g['services'].append({
                'slot': rec['slot'], 'prep': rec['prep'], 'waste': rec['waste'],
                'waste_ratio': rec['waste_ratio'], 'sheet': rec['sheet'],
            })
            g['total_prep'] += rec['prep']
            g['total_waste'] += rec['waste']

        # Build final rollup
        dishes = []
        for name, g in by_dish.items():
            n_services = len(g['services'])
            if g['total_prep'] <= 0:
                continue
            avg_waste_ratio = g['total_waste'] / g['total_prep']
            dishes.append({
                'name': name,
                'display_name': max(g['name_original_samples'], key=len),  # pick longest form
                'name_samples': sorted(g['name_original_samples']),
                'n_services': n_services,
                'total_prep': round(g['total_prep'], 2),
                'total_waste': round(g['total_waste'], 2),
                'avg_waste_ratio': round(avg_waste_ratio, 3),
                'avg_consumption_rate': round(1 - avg_waste_ratio, 3),
                'services': sorted(g['services'], key=lambda s: -s['waste_ratio']),
            })
        dishes.sort(key=lambda d: -d['avg_waste_ratio'])

        # --- Reporting ---
        self.stdout.write(self.style.HTTP_INFO(
            f'\n=== Production Tracker parse ({len(records)} clean services extracted) ==='))
        self.stdout.write(f'  Skipped (text-only prep/waste): {skipped_text_only}')
        self.stdout.write(f'  Skipped (no name / short name): {skipped_no_name}')
        self.stdout.write(f'  Unique dishes: {len(dishes)}')
        self.stdout.write(f'  Dishes with 3+ services: {sum(1 for d in dishes if d["n_services"] >= 3)}')

        repeat = [d for d in dishes if d['n_services'] >= 2]
        self.stdout.write(f'\n--- Top 10 high-waste dishes (2+ services) ---')
        for d in repeat[:10]:
            self.stdout.write(
                f'  waste={d["avg_waste_ratio"]:.1%}  eat={d["avg_consumption_rate"]:.1%}  '
                f'n={d["n_services"]}  {d["display_name"]}')

        self.stdout.write(f'\n--- Top 10 low-waste dishes (2+ services) ---')
        for d in list(reversed(repeat))[:10]:
            self.stdout.write(
                f'  waste={d["avg_waste_ratio"]:.1%}  eat={d["avg_consumption_rate"]:.1%}  '
                f'n={d["n_services"]}  {d["display_name"]}')

        if not opts['apply']:
            self.stdout.write(self.style.WARNING(
                '\nDry run — no JSON written. Re-run with --apply to save.'))
            return

        out_dir = Path(settings.BASE_DIR) / '.historical_stats'
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / 'production_tracker.json'
        out_path.write_text(json.dumps({
            'generated_at': datetime.utcnow().isoformat() + 'Z',
            'source': opts['xlsx'],
            'total_services': len(records),
            'unique_dishes': len(dishes),
            'dishes': dishes,
        }, indent=2))
        self.stdout.write(self.style.SUCCESS(
            f'\n✔ Wrote {out_path} ({out_path.stat().st_size} bytes).'))
