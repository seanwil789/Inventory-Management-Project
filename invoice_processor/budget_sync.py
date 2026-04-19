"""
Sync invoice totals from pipeline into the Wentworth budget spreadsheet.

Sources for invoice totals (checked in order):
  1. Invoice totals cache (.invoice_totals/) — populated by batch.py during processing
  2. OCR cache → re-parse invoice total from cached OCR text (no API cost)

Multi-page invoices (Sysco): uses the total from the LAST PAGE.

Usage:
  python budget_sync.py                        # sync current month
  python budget_sync.py --month 2026 4         # sync April 2026
  python budget_sync.py --month 2026 4 --dry-run
"""

import os
import sys
import re
import argparse
import tempfile
from datetime import date
from calendar import month_abbr

sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from drive import get_drive_client
from docai import ocr_with_docai
from parser import parse_invoice

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BUDGET_FILE = os.path.join(_PROJECT_ROOT, "Men's Wentworth Food Budget 2026.xlsx")
_INVOICE_TOTALS_DIR = os.path.join(_PROJECT_ROOT, ".invoice_totals")


def _check_totals_cache(year, month):
    """Check the invoice totals cache for pre-computed totals from batch.py."""
    import json
    cache_file = os.path.join(_INVOICE_TOTALS_DIR, f"{year}-{month:02d}.json")
    if os.path.exists(cache_file):
        with open(cache_file, "r") as f:
            return json.load(f)
    return []

# DB vendor name → budget sheet vendor name
VENDOR_MAP = {
    "Sysco":                        "Sysco",
    "Farm Art":                     "FarmArt",
    "Exceptional Foods":            "Exceptional",
    "Philadelphia Bakery Merchants": "PBM",
    "Delaware County Linen":        "Delaware Linen",
    "Colonial Village Meat Markets": "Colonial Meat Market",
    "Aramark":                      "Aramark",
}

# Month number → tab name in budget workbook
MONTH_TABS = {i: month_abbr[i] for i in range(1, 13)}

DATA_START_ROW = 10
_DRIVE_CLIENT = None


def _bootstrap_django():
    if not os.environ.get('DJANGO_SETTINGS_MODULE'):
        os.environ['DJANGO_SETTINGS_MODULE'] = 'myproject.settings'
        sys.path.insert(0, _PROJECT_ROOT)
        import django
        django.setup()


def _get_drive():
    global _DRIVE_CLIENT
    if _DRIVE_CLIENT is None:
        _DRIVE_CLIENT = get_drive_client()
    return _DRIVE_CLIENT


def _download_from_drive(filename: str) -> str | None:
    """Download a file from Drive by name. Returns local temp path or None."""
    client = _get_drive()
    query = f"name = '{filename}' and trashed = false"
    results = client.files().list(q=query, fields='files(id, name)').execute()
    files = results.get('files', [])
    if not files:
        return None
    content = client.files().get_media(fileId=files[0]['id']).execute()
    ext = os.path.splitext(filename)[1]
    tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
    tmp.write(content)
    tmp.close()
    return tmp.name


def _ocr_and_metadata(source_file: str, vendor: str) -> dict | None:
    """Download, OCR (cache hit), and extract metadata from a source file."""
    tmp_path = _download_from_drive(source_file)
    if not tmp_path:
        print(f"    [!] File not found in Drive: {source_file}")
        return None

    try:
        ocr_result = ocr_with_docai(tmp_path)
        if not ocr_result or not ocr_result.get("raw_text"):
            return None

        raw_text = ocr_result["raw_text"]
        parsed = parse_invoice(raw_text, vendor=vendor)

        # Extract Sysco-specific metadata for grouping
        sysco_meta = None
        if vendor == "Sysco":
            from parser import extract_sysco_metadata
            sysco_meta = extract_sysco_metadata(raw_text)

        page_total = parsed.get("invoice_total")
        items_sum = round(sum(
            it.get("extended_amount", it.get("unit_price", 0)) or 0
            for it in parsed.get("items", [])
        ), 2)
        is_last = "LAST PAGE" in raw_text.upper()

        return {
            "source_file": source_file,
            "raw_text": raw_text,
            "page_total": page_total,
            "items_sum": items_sum,
            "is_last_page": is_last,
            "sysco_meta": sysco_meta,
        }
    finally:
        os.remove(tmp_path)


def _get_invoice_total_from_files(vendor: str, source_files: list[str]) -> dict:
    """
    Download source files, OCR them (cache hit), parse invoice total.
    For non-Sysco vendors with single-page invoices.

    Returns: {"total": float|None, "method": str, "items_sum": float}
    """
    pages = []
    for source_file in sorted(source_files):
        page_data = _ocr_and_metadata(source_file, vendor)
        if page_data:
            pages.append(page_data)
    return _get_invoice_total_simple(pages)


def _get_invoice_total_simple(pages: list[dict]) -> dict:
    """
    Determine invoice total from pre-processed page data.
    Returns: {"total": float|None, "method": str, "items_sum": float}
    """
    best_total = None
    best_method = "not found"
    all_items_sum = sum(p["items_sum"] for p in pages)

    for page in pages:
        if page["page_total"] is not None:
            if page["is_last_page"]:
                best_total = page["page_total"]
                best_method = "last_page"
            elif best_method != "last_page":
                best_total = page["page_total"]
                best_method = "parsed"

    if best_total is None and all_items_sum > 0:
        best_total = all_items_sum
        best_method = "items_sum"

    return {
        "total": best_total,
        "method": best_method,
        "items_sum": round(all_items_sum, 2),
    }


def _get_invoices_for_month(year: int, month: int) -> list[dict]:
    """
    Get all invoices for a month: (vendor, date, source_files).

    For Sysco: groups by INVOICE NUMBER (from OCR metadata) and uses
    DELV. DATE for accurate delivery dating. This correctly separates
    multi-page invoices even when photographed together.

    For other vendors: groups by (vendor, date) from DB.
    """
    _bootstrap_django()
    from myapp.models import InvoiceLineItem

    # Get distinct (vendor, date, source_file) combinations
    combos = (
        InvoiceLineItem.objects
        .filter(
            invoice_date__year=year,
            invoice_date__month=month,
            vendor__isnull=False,
        )
        .exclude(source_file='')
        .values('vendor__name', 'invoice_date', 'source_file')
        .distinct()
        .order_by('invoice_date', 'vendor__name', 'source_file')
    )

    # Separate Sysco files from others
    sysco_files = []
    other_invoices = {}

    for c in combos:
        if c['vendor__name'] == 'Sysco':
            sysco_files.append(c['source_file'])
        else:
            key = (c['vendor__name'], c['invoice_date'])
            if key not in other_invoices:
                other_invoices[key] = {
                    "vendor_db": c['vendor__name'],
                    "vendor_budget": VENDOR_MAP.get(c['vendor__name'], c['vendor__name']),
                    "date": c['invoice_date'],
                    "source_files": [],
                }
            other_invoices[key]["source_files"].append(c['source_file'])

    # Check invoice totals cache — use cached totals when available
    cached_totals = _check_totals_cache(year, month)
    cached_lookup = {}  # (vendor, date) → total
    for ct in cached_totals:
        key = (ct['vendor'], ct['date'])
        if key not in cached_lookup:
            cached_lookup[key] = ct['total']
        else:
            # Multiple entries for same vendor+date = multi-page, use the latest
            cached_lookup[key] = ct['total']

    # Attach cached totals to non-Sysco invoices
    result = []
    for inv in other_invoices.values():
        cache_key = (inv['vendor_db'], str(inv['date']))
        if cache_key in cached_lookup:
            inv['_precomputed_total'] = {
                'total': cached_lookup[cache_key],
                'method': 'batch_cache',
                'items_sum': 0,
            }
        result.append(inv)

    if cached_lookup:
        cached_count = sum(1 for inv in result if '_precomputed_total' in inv)
        print(f"  {cached_count} invoices found in totals cache (fast path)")

    # Group Sysco files by invoice number using OCR metadata
    if sysco_files:
        print(f"  Grouping {len(sysco_files)} Sysco pages by invoice number...")
        sysco_pages = []  # list of page metadata dicts
        for sf in sorted(set(sysco_files)):
            page_data = _ocr_and_metadata(sf, "Sysco")
            if page_data:
                sysco_pages.append(page_data)

        # Group by invoice_number (primary) or manifest (fallback)
        invoice_groups = {}
        for page in sysco_pages:
            meta = page.get("sysco_meta") or {}
            inv_num = meta.get("invoice_number") or meta.get("manifest") or "unknown"
            if inv_num not in invoice_groups:
                invoice_groups[inv_num] = {
                    "pages": [],
                    "delivery_date": None,
                }
            invoice_groups[inv_num]["pages"].append(page)
            # Use delivery date from whichever page has it
            if meta.get("delivery_date") and not invoice_groups[inv_num]["delivery_date"]:
                invoice_groups[inv_num]["delivery_date"] = meta["delivery_date"]

        # Convert to invoice entries
        from datetime import datetime
        for inv_num, group in invoice_groups.items():
            # Parse delivery date
            delv_date = None
            if group["delivery_date"]:
                try:
                    delv_date = datetime.strptime(group["delivery_date"], "%m/%d/%y").date()
                except ValueError:
                    try:
                        delv_date = datetime.strptime(group["delivery_date"], "%m/%d/%Y").date()
                    except ValueError:
                        pass

            # Get invoice total from grouped pages
            total_result = _get_invoice_total_simple(group["pages"])

            if delv_date is None:
                print(f"    [!] Sysco invoice {inv_num}: no delivery date found — flagging for review")

            result.append({
                "vendor_db": "Sysco",
                "vendor_budget": "Sysco",
                "date": delv_date,
                "source_files": [p["source_file"] for p in group["pages"]],
                "_precomputed_total": total_result,
                "_invoice_number": inv_num,
                "_page_count": len(group["pages"]),
            })

            print(f"    Invoice {inv_num}: {len(group['pages'])} pages, "
                  f"delv={group['delivery_date'] or '?'}, "
                  f"total=${total_result['total']:,.2f}" if total_result['total'] else
                  f"    Invoice {inv_num}: {len(group['pages'])} pages, "
                  f"delv={group['delivery_date'] or '?'}, total=?")

    return result

def _find_next_empty_row(ws, start_row=DATA_START_ROW) -> int:
    row = start_row
    while ws[f"B{row}"].value is not None:
        row += 1
    return row


def _get_existing_entries(ws, start_row=DATA_START_ROW) -> set:
    existing = set()
    row = start_row
    while ws[f"B{row}"].value is not None:
        d = ws[f"B{row}"].value
        s = ws[f"C{row}"].value or ""
        if d:
            if hasattr(d, 'date'):
                d = d.date()
            existing.add((d, s.strip().lower()))
        row += 1
    return existing


def sync_month(year: int, month: int, dry_run: bool = False) -> dict:
    tab_name = MONTH_TABS.get(month)
    if not tab_name:
        print(f"[!] Invalid month: {month}")
        return {"added": 0, "skipped": 0, "failed": 0}

    invoices = _get_invoices_for_month(year, month)
    if not invoices:
        print(f"  No invoice data for {tab_name} {year}")
        return {"added": 0, "skipped": 0, "failed": 0}

    print(f"  {len(invoices)} invoices for {tab_name} {year}")

    if not os.path.exists(BUDGET_FILE):
        print(f"[!] Budget file not found: {BUDGET_FILE}")
        return {"added": 0, "skipped": 0, "failed": 0}

    wb = openpyxl.load_workbook(BUDGET_FILE)
    if tab_name not in wb.sheetnames:
        print(f"[!] Tab '{tab_name}' not found")
        return {"added": 0, "skipped": 0, "failed": 0}

    ws = wb[tab_name]
    existing = _get_existing_entries(ws)
    next_row = _find_next_empty_row(ws)

    added = 0
    skipped = 0
    failed = 0

    for inv in invoices:
        # Skip invoices with no date (Sysco pages with unreadable delivery date)
        if inv["date"] is None:
            inv_num = inv.get("_invoice_number", "?")
            print(f"  [!] Sysco invoice {inv_num}: no delivery date — needs manual entry")
            failed += 1
            continue

        key = (inv["date"], inv["vendor_budget"].strip().lower())
        if key in existing:
            print(f"  [skip] {inv['date']} {inv['vendor_budget']} — already exists")
            skipped += 1
            continue

        # Sysco invoices have precomputed totals from the grouping step
        if "_precomputed_total" in inv:
            result = inv["_precomputed_total"]
            pages_label = f"{inv.get('_page_count', '?')} pages, inv#{inv.get('_invoice_number', '?')}"
            print(f"  Processing {inv['date']} Sysco ({pages_label})...")
        else:
            print(f"  Processing {inv['date']} {inv['vendor_db']} ({len(inv['source_files'])} pages)...")
            result = _get_invoice_total_from_files(inv["vendor_db"], inv["source_files"])

        if result["total"] is None:
            print(f"    [!] Could not determine invoice total — skipping")
            failed += 1
            continue

        total = round(result["total"], 2)
        method = result["method"]

        # Validation: compare parsed total to items sum
        if result["items_sum"] > 0 and method != "items_sum":
            gap = abs(total - result["items_sum"])
            pct = (gap / total * 100) if total > 0 else 0
            if pct > 20:
                print(f"    [!] Large gap: total=${total:,.2f} vs items=${result['items_sum']:,.2f} "
                      f"({pct:.0f}% — may be missing parsed items)")

        if dry_run:
            print(f"    [DRY RUN] Row {next_row}: {inv['date']}  {inv['vendor_budget']:<25}  "
                  f"${total:,.2f}  ({method})")
        else:
            ws[f"B{next_row}"] = inv["date"]
            ws[f"C{next_row}"] = inv["vendor_budget"]
            ws[f"D{next_row}"] = total
            ws[f"B{next_row}"].number_format = "M/D/YYYY"
            ws[f"D{next_row}"].number_format = '#,##0.00'
            print(f"    [+] Row {next_row}: {inv['date']}  {inv['vendor_budget']:<25}  "
                  f"${total:,.2f}  ({method})")

        next_row += 1
        added += 1

    if not dry_run and added > 0:
        wb.save(BUDGET_FILE)
        print(f"\n  [saved] {BUDGET_FILE}")

    return {"added": added, "skipped": skipped, "failed": failed}


def main():
    parser = argparse.ArgumentParser(description="Sync invoice totals to budget sheet")
    parser.add_argument("--month", nargs=2, metavar=("YEAR", "MONTH"),
                        help="Year and month to sync (e.g., 2026 4)")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing")
    args = parser.parse_args()

    if args.month:
        year, month = int(args.month[0]), int(args.month[1])
    else:
        today = date.today()
        year, month = today.year, today.month

    result = sync_month(year, month, dry_run=args.dry_run)
    print(f"\n  Added: {result['added']}, Skipped: {result['skipped']}, Failed: {result['failed']}")


if __name__ == "__main__":
    main()
